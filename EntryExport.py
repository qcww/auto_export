# coding=utf-8
'''
Created on 2016-8-16
@author: Jennifer
Project:使用chrome浏览器，安装chromewebdriver.exe
'''
from selenium import webdriver 
from selenium.common.exceptions import SessionNotCreatedException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException
import os
import sys
import time
import urllib3
from time import sleep
import configparser
import requests
import locale
import openpyxl
import json
import calendar
from urllib import parse
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait

import regedit

class fpdk:
    def __init__(self):
        self.config_file = "./config/config.ini"
        if os.path.exists(self.config_file) == False:
            self.config_file = regedit.get_client_path()+"\\config\\config.ini"

        self.config = configparser.ConfigParser()
        self.config.read(self.config_file,encoding='utf-8')
        locale.setlocale(locale.LC_ALL,'en')
        locale.setlocale(locale.LC_CTYPE,'chinese')

    def check_driver(self):
        base_dir=os.getcwd()
        sys.path.append(base_dir)
        options=webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        try:
            self.driver=webdriver.Chrome(options=options)
        except SessionNotCreatedException:
            return '该版本谷歌浏览器与软件不兼容，如有问题请联系客服'
        except WebDriverException:
            return "请安装谷歌浏览器后重试"
        # else:
        #     return '请下载指定版本谷歌浏览器后重试'
        
        self.Action = ActionChains(self.driver)
        try:
            login_err = self.invoice_login()
            if login_err != '':
                return login_err
        except :
            return "登录发票验证网站失败"
        urllib3.disable_warnings()
        self.set_driver_cookie()
        return ''

    # 登录税务局网站
    def invoice_login(self):
        self.driver.get(self.config['link']['entry_host'])
        self.fpdk_version = self.get_by_driver("ymbb")
        
        self.driver.execute_script("$('#okButton').click();")
        sleep(3)
        passwd = self.driver.find_element_by_id("password1")
        login_btn = self.driver.find_element_by_id("submit")
        time.sleep(2)

        passwd.clear()
        cert = self.config['client']['cert']
        self.driver.execute_script("$('#password').val('%s');" % cert)
        # self.driver.execute_script("$('#password1').val('12345678');")
        login_btn_click = self.Action.click(login_btn)
        login_btn_click.perform()

        sleep(2)
        # 判断是否有密码错误的提示框出来
        try:
            pass_err = self.driver.find_element_by_class_name("theme-popover")
            if pass_err:
                err_msg = self.driver.find_element_by_class_name("theme-popover").text
                if '口令验证失败' in err_msg:
                    return "证书口令不正确，请联系客服修改"
                return err_msg.split("\n")[-1]
        except:
            pass

        return ''

    # 抓取进项票
    def scrap_data(self,ymd):
        # 添加延迟防止token等未生成
        time.sleep(2)
        self.export_no_rz(ymd)
        ret = self.export_rz(ymd)
        self.driver.quit()
        return ret

    # 导出未认证进项发票
    def export_no_rz(self,ymd):
        # 固定配置参数
        _,e = calendar.monthrange(int(ymd[0:4]),int(ymd[5:7]))
        rq_q = ymd
        rq_z = ymd[0:8] + str(e)
        entry_fix_0 = self.config['export']['entry_fix_0']
        payload = "id=dkgxquery&fpdm=&fphm=&rq_q=%s&rq_z=%s&xfsbh=&rzzt=0&glzt=0&fpzt=0&fplx=-1&cert=%s&token=%s&ymbb=%s&aoData=%s" % (rq_q,rq_z,self.get_by_driver('cert'),self.get_by_driver('token'),self.get_by_driver('ymbb'),entry_fix_0.replace(':','%'))
        res = self.driver_post(self.config['link']['entry_host'] + self.config['link']['dkgx_url'],payload)
        if 'key3' not in res:
            return {"code":500,"text":"操作失败，所选时间范围无效"}
        else:
            if len(res['key3']['aaData']) == 0:
                return {"code":404,"text":"所选时间范围无数据"}
            invoice = []
            for d in res['key3']['aaData']:
                temp = {"dedicated_status":1,"code":d[1],"number":d[2],"open_date":d[3],"sale_name":d[4],"money":d[5],"tax_amount":d[6]}
                invoice.append(temp)
            post_data = {"credit_code":self.get_by_driver('cert'),"invoice":json.dumps(invoice)}
      
            ret = self.http_post(self.config['link']['host'] + self.config['upload']['purchas_verify'],post_data)
            return ret

    # 导出已认证进项发票
    def export_rz(self,ymd):
        tjyf = ymd[0:-3].replace('-','')
        cert = self.get_by_driver('cert')
        entry_fix_1 = self.config['export']['entry_fix_1']
        payload = "tjyf=%s&id=dkmx&cert=%s&token=%s&fpdm=&fphm=&xfsbh=&qrrzrq_q=&qrrzrq_z=&fply=0&ymbb=%s&qt=wq&aoData=%s" % (tjyf,cert,self.get_by_driver('token'),self.get_by_driver('ymbb'),entry_fix_1.replace(':','%'))
        res = self.driver_post(self.config['link']['entry_host'] + self.config['link']['dktj_url'],payload)
        if 'key3' not in res:
            return {"code":404,"text":res['key2']}
        else:
            # print(tjyf,"已勾选数据",res['key3']['aaData'])
            if len(res['key3']['aaData']) > 0:
                return self.save_purchas_file(cert,tjyf,res['key3']['aaData'])
            return {"code":404,"text":"未找到历史数据"}


    # 生成并上传数据
    def save_purchas_file(self,credit_code,period,ret_json):
        sal_xls = '进项发票_%s_%s.xlsx' % (credit_code,period)
        e_cols = {"A":0,"B":1,"C":2,"D":3,"E":"","F":4,"G":5,"H":6,"I":7,"J":8,"K":9,"L":10,"M":11,"N":12}

        wb = openpyxl.load_workbook(regedit.get_client_path()+'\\template\\temp.xlsx')
        ws = wb['1']
        row_start = 4
        ws['B2'] = credit_code
        ws['H2'] = period
        for em in ret_json:
            # em_list.append(em['number'])
            for rw in e_cols:
                key_name = e_cols[rw]
                if key_name != '':
                    ws[rw+str(row_start)] = em[key_name]
            row_start += 1
        wb.save(regedit.get_client_path()+'\\exp_file\\'+sal_xls)
        return {"code":200,"text":"进项票勾选文件保存成功","p_file":sal_xls}



    # 通过driver 获取js变量
    def get_by_driver(self,val):
        return self.driver.execute_script("return %s" % val)


    # 通过driver的cookie 发送post数据
    def driver_post(self,url,data={}):
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            # 'Host': 'etax.anhui.chinatax.gov.cn', 
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
            'Cookie':self._cookie
        }
        ret = requests.post(url,data,headers=headers,verify=False)
        return ret.json()

    # 通过requests 发送post数据
    def http_post(self,url,data={}):
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
        }
        ret = requests.post(url,data,headers=headers,verify=False)
        return ret.json()

    # 设置request 的 cookie
    def set_driver_cookie(self):
        cookies = self.driver.get_cookies()
        self._cookie = ''
        for item in cookies:
            self._cookie = self._cookie + item['name']+'='+item['value']+';'
   

# fk = fpdk()
# fk.scrap_data()

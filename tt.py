# -*- coding: utf-8 -*- 

import urllib.parse
import sys
import html
import re
import os
import cgi
import urllib.parse
import calendar
import requests
import json
import openpyxl
import regedit

# def unescape(string):
#     string = urllib.parse.unquote(string)
#     quoted = html.unescape(string).encode(sys.getfilesystemencoding()).decode('utf-8')
#     #转成中文
#     return re.sub(r'%u([a-fA-F0-9]{4}|[a-fA-F0-9]{2})', lambda m: chr(int(m.group(1), 16)), quoted)
# cc = "%5B%7B%22name%22%3A%22sEcho%22%2C%22value%22%3A1%7D%2C%7B%22name%22%3A%22iColumns%22%2C%22value%22%3A14%7D%2C%7B%22name%22%3A%22sColumns%22%2C%22value%22%3A%22%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%22%7D%2C%7B%22name%22%3A%22iDisplayStart%22%2C%22value%22%3A0%7D%2C%7B%22name%22%3A%22iDisplayLength%22%2C%22value%22%3A50%7D%2C%7B%22name%22%3A%22mDataProp_0%22%2C%22value%22%3A0%7D%2C%7B%22name%22%3A%22mDataProp_1%22%2C%22value%22%3A1%7D%2C%7B%22name%22%3A%22mDataProp_2%22%2C%22value%22%3A2%7D%2C%7B%22name%22%3A%22mDataProp_3%22%2C%22value%22%3A3%7D%2C%7B%22name%22%3A%22mDataProp_4%22%2C%22value%22%3A4%7D%2C%7B%22name%22%3A%22mDataProp_5%22%2C%22value%22%3A5%7D%2C%7B%22name%22%3A%22mDataProp_6%22%2C%22value%22%3A6%7D%2C%7B%22name%22%3A%22mDataProp_7%22%2C%22value%22%3A7%7D%2C%7B%22name%22%3A%22mDataProp_8%22%2C%22value%22%3A8%7D%2C%7B%22name%22%3A%22mDataProp_9%22%2C%22value%22%3A9%7D%2C%7B%22name%22%3A%22mDataProp_10%22%2C%22value%22%3A10%7D%2C%7B%22name%22%3A%22mDataProp_11%22%2C%22value%22%3A11%7D%2C%7B%22name%22%3A%22mDataProp_12%22%2C%22value%22%3A12%7D%2C%7B%22name%22%3A%22mDataProp_13%22%2C%22value%22%3A13%7D%5D"
# print(cc.replace('%',':'))

# config_data = "rzzt=0&glzt=0&fpzt=0&fplx=01&xfsbh=&id=dkgxquery&fpdm=&fphm=&aoData=%5B%7B%22name%22%3A%22sEcho%22%2C%22value%22%3A1%7D%2C%7B%22name%22%3A%22iColumns%22%2C%22value%22%3A15%7D%2C%7B%22name%22%3A%22sColumns%22%2C%22value%22%3A%22%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%22%7D%2C%7B%22name%22%3A%22iDisplayStart%22%2C%22value%22%3A0%7D%2C%7B%22name%22%3A%22iDisplayLength%22%2C%22value%22%3A1000000%7D%2C%7B%22name%22%3A%22mDataProp_0%22%2C%22value%22%3A0%7D%2C%7B%22name%22%3A%22mDataProp_1%22%2C%22value%22%3A1%7D%2C%7B%22name%22%3A%22mDataProp_2%22%2C%22value%22%3A2%7D%2C%7B%22name%22%3A%22mDataProp_3%22%2C%22value%22%3A3%7D%2C%7B%22name%22%3A%22mDataProp_4%22%2C%22value%22%3A4%7D%2C%7B%22name%22%3A%22mDataProp_5%22%2C%22value%22%3A5%7D%2C%7B%22name%22%3A%22mDataProp_6%22%2C%22value%22%3A6%7D%2C%7B%22name%22%3A%22mDataProp_7%22%2C%22value%22%3A7%7D%2C%7B%22name%22%3A%22mDataProp_8%22%2C%22value%22%3A8%7D%2C%7B%22name%22%3A%22mDataProp_9%22%2C%22value%22%3A9%7D%2C%7B%22name%22%3A%22mDataProp_10%22%2C%22value%22%3A10%7D%2C%7B%22name%22%3A%22mDataProp_11%22%2C%22value%22%3A11%7D%2C%7B%22name%22%3A%22mDataProp_12%22%2C%22value%22%3A12%7D%2C%7B%22name%22%3A%22mDataProp_13%22%2C%22value%22%3A13%7D%2C%7B%22name%22%3A%22mDataProp_14%22%2C%22value%22%3A14%7D%5D"
# par = urllib.parse.parse_qsl(config_data)
# print(dict(par))

# ymd = '2020-12-02'
# print(ymd[0:-3].replace('-',''))

# print(ymd[0:4],ymd[5:7])
# _,e = calendar.monthrange(int(ymd[0:4]),int(ymd[5:7]))
# print(ymd[0:8] + '01')
# print(ymd[0:8] + str(e))
# invoice_data = [{"dedicated_status": 1, "code": "3100193130", "number": "03291721", "open_date": "2019-11-18", "sale_name": "\u4e0a\u6d77\u5c82\u5b81\u4f20\u52a8\u673a\u68b0\u6709\u9650\u516c\u53f8", "money": "14601.77", "tax_amount": "1898.23"}, {"dedicated_status": 1, "code": "1100191130", "number": "07567351", "open_date": "2019-11-11", "sale_name": "\u4e07\u6cca\u4e3d\u666f\uff08\u5317\u4eac\uff09\u6587\u5316\u65c5\u6e38\u5f00\u53d1\u6709\u9650\u516c\u53f8", "money": "845.28", "tax_amount": "50.72"}]
# dd = {"credit_code": "91340100MA2NPN203B", "invoice": json.dumps(invoice_data)}
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
# }
# ret = requests.post('http://tinterface.hfxscw.com/interface.php?r=tax/handle-add-income',data=dd,headers=headers)
# print(ret.text)
ret_json = [['1', '3700191130', '27295617', '2019-11-24', '临清市格达轴承有限公司', '99111.51', '12884.49', '12884.49', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295617','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2icP4CAxBKiBtcFXJZ45YJg%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['2', '4100192130', '04589453', '2019-11-25', '安阳博恒轴承制造有限公司', '97435.39', '12666.61', '12666.61', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('4100192130','04589453','2019-11-25','01','edQPhY2ugYIRPTkKzVIGmZnaUiphmwmX9FNChtgFSkc%3D','安阳博恒轴承制造有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['3', '3700191130', '27295618', '2019-11-24', '临清市格达轴承有限公司', '96064.61', '12488.39', '12488.39', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295618','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2g0Q2XGZmcULCtPI4zCQXqs%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['4', '3700191130', '27295619', '2019-11-24', '临清市格达轴承有限公司', '97752.21', '12707.79', '12707.79', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295619','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2t6d4tSy2F4c8q%2Fs3HQV3w0%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['5', '3700191130', '27295620', '2019-11-24', '临清市格达轴承有限公司', '97752.21', '12707.79', '12707.79', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295620','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2p9qvEmx4oC8iCL3vTRqByI%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['6', '3700191130', '27295621', '2019-11-24', '临清市格达轴承有限公司', '98831.86', '12848.14', '12848.14', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295621','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2pZgxfodtsd%2BIla0BwRi0yc%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['7', '3700191130', '27295622', '2019-11-24', '临清市格达轴承有限公司', '98783.18', '12841.82', '12841.82', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295622','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2nJ8eAZ8bltI2qPS5mnjWVQ%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['8', '3700191130', '27295623', '2019-11-24', '临清市格达轴承有限公司', '97977.87', '12737.13', '12737.13', '2019-11-27', '增值税专票', '抵扣', '正常', '正常', "<a onclick=cxFpmx('3700191130','27295623','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2hM3sPVUJ%2FIWybdmFzt3w7A%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"], ['9', '3700191130', '27295624', '2019-11-24', '临清市格达轴承有限公司', '85129.2', '11066.8', '11066.8', '2019-11-27', '增值税专票', '抵扣','正常', '正常', "<a onclick=cxFpmx('3700191130','27295624','2019-11-24','01','W4%2FElzcTUQ1KI0g06OwU2gMhPzUQt7Ufk2tDo%2FWLyos%3D','临清市格达轴承有限公司','0'); href='javascript:void(0);'><font color=red>查看明细信息</a>"]]

credit_code = '91340100MA2NPN203B'
period = '201911'
sal_xls = regedit.get_client_path()+'\\exp_file\\%s-%s.xlsx' % (credit_code,period)
e_cols = {"A":0,"B":1,"C":2,"D":3,"E":"","F":4,"G":5,"H":6,"I":7,"J":8,"K":9,"L":10,"M":11,"N":12}

wb = openpyxl.load_workbook(regedit.get_client_path()+'\\template\\temp.xlsx')
ws = wb['1']
row_start = 4
ws['B2'] = credit_code
ws['H2'] = period
for em in ret_json:
    # em_list.append(em['number'])
    for rw in e_cols:
        key_name = e_cols[rw]
        if key_name != '':
            ws[rw+str(row_start)] = em[key_name]
    row_start += 1
wb.save(sal_xls)

file_handle = open(sal_xls, 'rb')

headers={
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'} 
file = {'upfile':(sal_xls, file_handle)}
post_data = {"period":period,"credit_code":credit_code}
r = requests.post(url = "http://tinterface.hfxscw.com/interface.php?r=tax/import-purchase-invoice",data=post_data, headers = headers,files = file)
file_handle.close()
try:
    os.remove(sal_xls)
except(FileNotFoundError):
    print("文件不存在")
print(r.text)